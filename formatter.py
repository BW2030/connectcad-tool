"""Generate ConnectCAD-compatible worksheet (CSV + XLSX).

VW 2026 / ConnectCAD 2026:
  - Device definitions are embedded in 3D symbols (Resource Manager)
  - 'Create Devices from Worksheet' matches Make+Model against symbols
  - Required columns: Make, Model, Qty, Name  (+ optional Room, Rack, Rack U)
  - Best import formats: CSV or Excel (.xlsx)
"""
import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ConnectCAD 2026 worksheet columns
COLUMNS = ["Make", "Model", "Qty", "Name", "Room", "Rack", "Rack U"]


def build_rows(device: dict, qty: int = 1, room: str = "", rack: str = "", rack_u: str = "") -> list[list[str]]:
    """Build ConnectCAD 2026 worksheet rows (header + data)."""
    name_tag = f"{device.get('make', '')} {device.get('model', '')}".strip()
    header = COLUMNS
    data_row = [
        device.get("make", ""),
        device.get("model", ""),
        str(qty),
        name_tag,
        room,
        rack,
        rack_u,
    ]
    return [header, data_row]


def to_csv(rows: list[list[str]]) -> str:
    """Export rows as CSV (UTF-8 with BOM for Excel compatibility)."""
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=",", quoting=csv.QUOTE_MINIMAL)
    for row in rows:
        writer.writerow(row)
    return "\ufeff" + buf.getvalue()  # BOM → Excel öffnet UTF-8 korrekt


def to_tab_delimited(rows: list[list[str]]) -> str:
    """Legacy: tab-delimited (für ältere VW-Versionen)."""
    return "\n".join("\t".join(row) for row in rows)


def to_excel_bytes(rows: list[list[str]]) -> bytes:
    """Export rows as an Excel workbook (.xlsx) and return bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ConnectCAD Import"

    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri")
    data_font = Font(name="Calibri")

    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.font = data_font

    # Auto-size columns (approximate)
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
